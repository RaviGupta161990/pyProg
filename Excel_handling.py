from openpyxl import Workbook
from openpyxl.utils import get_column_letter


wb = Workbook()
ws = wb.active
ws.title = "Sample 1"

for row in range(1,40):
    ws.append(range(600))

ws2 = wb.create_sheet(title="Sample 2")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Sample 3")

for row in range(10,40):
    for col in range(27,54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

wb.save(filename="SampleWB.xlsx")
